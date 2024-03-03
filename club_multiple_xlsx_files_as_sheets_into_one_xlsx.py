import os
from openpyxl import Workbook
from openpyxl import load_workbook


def club_files():
    dir_containing_files = "C:\\Users\\XXXX\\Desktop\\New folder"

    dest_wb = Workbook()
    for root, dir, filenames in os.walk(dir_containing_files):
        for file in filenames:
            print('something hoing!')
            if '.xlsx' in file:
                file_name = file.split('.')[0]
                # Absolute Path for Excel files
                file_path = os.path.abspath(os.path.join(root, file))

                # Create new sheet in destination Workbook
                dest_wb.create_sheet(file_name)
                dest_ws = dest_wb[file_name]

                # =====New Code====#

                # Read source data
                source_wb = load_workbook(file_path)
                source_sheet = source_wb.active
                for row in source_sheet.rows:
                    for cell in row:
                        dest_ws[cell.coordinate] = cell.value
                # =================#

    dest_wb.save("file_16_03_2023.xlsx")


club_files()
